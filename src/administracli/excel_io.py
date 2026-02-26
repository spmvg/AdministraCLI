"""
Read and write the AdministraCLI Excel workbook.
"""

import uuid
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Optional

import openpyxl
import pandas as pd

from administracli.models import (
    Administracli,
    IncomingInvoice,
    OutgoingInvoice,
    Transaction,
    VATDeclaration,
)

TRANSACTIONS_SHEET = "transactions"
INCOMING_INVOICES_SHEET = "incoming_invoices"
OUTGOING_INVOICES_SHEET = "outgoing_invoices"
VAT_DECLARATIONS_SHEET = "vat_declarations"
SETTINGS_SHEET = "settings"

TRANSACTION_COLUMNS = [
    "Date",
    "Amount",
    "Bank Account",
    "Description",
    "_id",
    "_incoming_invoice_id",
    "_outgoing_invoice_id",
    "_vat_declaration_id",
    "_category",
]

INCOMING_INVOICE_COLUMNS = [
    "Date",
    "Amount",
    "Counterparty",
    "vat_rate",
    "vat_rate_abroad_from_outside_eu",
    "vat_rate_abroad_from_inside_eu",
    "_id",
]

OUTGOING_INVOICE_COLUMNS = [
    "Date",
    "Amount",
    "Counterparty",
    "vat_rate",
    "_id",
]

VAT_DECLARATION_COLUMNS = [
    "period_start_date_inclusive",
    "period_end_date_exclusive",
    "_revenue_ex_vat",
    "_revenue_vat",
    "_reverse_charge_outside_eu_ex_vat",
    "_reverse_charge_outside_eu_vat",
    "_reverse_charge_inside_eu_ex_vat",
    "_reverse_charge_inside_eu_vat",
    "_input_vat",
    "_id",
]


def _generate_id() -> str:
    return str(uuid.uuid4())


def _to_optional_str(value) -> Optional[str]:
    if pd.isna(value):
        return None
    return str(value)


def _to_optional_decimal(value) -> Optional[Decimal]:
    if pd.isna(value):
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


def init_workbook(path: str) -> None:
    """Create a new Excel workbook with empty worksheets and correct headers."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    ws = wb.active
    wb.remove(ws)

    for sheet_name, columns in [
        (TRANSACTIONS_SHEET, TRANSACTION_COLUMNS),
        (INCOMING_INVOICES_SHEET, INCOMING_INVOICE_COLUMNS),
        (OUTGOING_INVOICES_SHEET, OUTGOING_INVOICE_COLUMNS),
        (VAT_DECLARATIONS_SHEET, VAT_DECLARATION_COLUMNS),
    ]:
        ws = wb.create_sheet(title=sheet_name)
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=1, column=col_idx, value=col_name)

    # Settings sheet (key-value pairs)
    ws = wb.create_sheet(title=SETTINGS_SHEET)
    ws.cell(row=1, column=1, value="Key")
    ws.cell(row=1, column=2, value="Value")
    ws.cell(row=2, column=1, value="cit_amount")
    ws.cell(row=2, column=2, value=None)

    wb.save(path)


def load_workbook(path: str) -> Administracli:
    """Load an AdministraCLI workbook into dataclass instances."""
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")

    transactions = _load_transactions(path)
    incoming_invoices = _load_incoming_invoices(path)
    outgoing_invoices = _load_outgoing_invoices(path)
    vat_declarations = _load_vat_declarations(path)
    cit_amount = _load_cit_amount(path)

    data = Administracli(
        transactions=transactions,
        incoming_invoices=incoming_invoices,
        outgoing_invoices=outgoing_invoices,
        vat_declarations=vat_declarations,
        cit_amount=cit_amount,
    )
    _compute_vat_fields(data)
    return data


def _read_sheet(path: str, sheet_name: str, expected_columns: list[str]) -> pd.DataFrame:
    """Read a sheet, returning an empty DataFrame with expected columns if the sheet is empty."""
    try:
        df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    except ValueError:
        # Sheet doesn't exist
        df = pd.DataFrame(columns=expected_columns)

    for col in expected_columns:
        if col not in df.columns:
            df[col] = None

    return df


def _load_cit_amount(path: str) -> Optional[Decimal]:
    """Load the definitive corporate income tax amount from the settings sheet."""
    try:
        df = pd.read_excel(path, sheet_name=SETTINGS_SHEET, engine="openpyxl")
    except ValueError:
        return None
    row = df.loc[df["Key"] == "cit_amount"]
    if row.empty:
        return None
    val = row.iloc[0]["Value"]
    if pd.isna(val):
        return None
    try:
        return Decimal(str(val))
    except InvalidOperation:
        return None


def _load_transactions(path: str) -> list[Transaction]:
    df = _read_sheet(path, TRANSACTIONS_SHEET, TRANSACTION_COLUMNS)
    transactions = []
    for _, row in df.iterrows():
        _id = _to_optional_str(row.get("_id"))
        if _id is None:
            _id = _generate_id()

        transactions.append(
            Transaction(
                date=row["Date"] if not pd.isna(row.get("Date")) else None,
                amount=Decimal(str(row["Amount"])) if not pd.isna(row.get("Amount")) else Decimal(0),
                bank_account=str(row["Bank Account"]) if not pd.isna(row.get("Bank Account")) else "",
                description=_to_optional_str(row.get("Description")),
                _id=_id,
                _incoming_invoice_id=_to_optional_str(row.get("_incoming_invoice_id")),
                _outgoing_invoice_id=_to_optional_str(row.get("_outgoing_invoice_id")),
                _vat_declaration_id=_to_optional_str(row.get("_vat_declaration_id")),
                _category=_to_optional_str(row.get("_category")),
            )
        )
    return transactions


def _load_incoming_invoices(path: str) -> list[IncomingInvoice]:
    df = _read_sheet(path, INCOMING_INVOICES_SHEET, INCOMING_INVOICE_COLUMNS)
    invoices = []
    for _, row in df.iterrows():
        _id = _to_optional_str(row.get("_id"))
        if _id is None:
            _id = _generate_id()

        invoices.append(
            IncomingInvoice(
                date=row["Date"] if not pd.isna(row.get("Date")) else None,
                amount=Decimal(str(row["Amount"])) if not pd.isna(row.get("Amount")) else Decimal(0),
                counterparty=str(row["Counterparty"]) if not pd.isna(row.get("Counterparty")) else "",
                _id=_id,
                vat_rate=_to_optional_decimal(row.get("vat_rate")),
                vat_rate_abroad_from_outside_eu=_to_optional_decimal(row.get("vat_rate_abroad_from_outside_eu")),
                vat_rate_abroad_from_inside_eu=_to_optional_decimal(row.get("vat_rate_abroad_from_inside_eu")),
            )
        )
    return invoices


def _load_outgoing_invoices(path: str) -> list[OutgoingInvoice]:
    df = _read_sheet(path, OUTGOING_INVOICES_SHEET, OUTGOING_INVOICE_COLUMNS)
    invoices = []
    for _, row in df.iterrows():
        _id = _to_optional_str(row.get("_id"))
        if _id is None:
            _id = _generate_id()

        vat_rate = _to_optional_decimal(row.get("vat_rate"))
        if vat_rate is None:
            vat_rate = Decimal(0)

        invoices.append(
            OutgoingInvoice(
                date=row["Date"] if not pd.isna(row.get("Date")) else None,
                amount=Decimal(str(row["Amount"])) if not pd.isna(row.get("Amount")) else Decimal(0),
                counterparty=str(row["Counterparty"]) if not pd.isna(row.get("Counterparty")) else "",
                vat_rate=vat_rate,
                _id=_id,
            )
        )
    return invoices


def _load_vat_declarations(path: str) -> list[VATDeclaration]:
    df = _read_sheet(path, VAT_DECLARATIONS_SHEET, VAT_DECLARATION_COLUMNS)
    declarations = []
    for _, row in df.iterrows():
        _id = _to_optional_str(row.get("_id"))
        if _id is None:
            _id = _generate_id()

        start = row.get("period_start_date_inclusive")
        end = row.get("period_end_date_exclusive")
        if pd.isna(start) or pd.isna(end):
            continue  # skip rows without valid period

        declarations.append(
            VATDeclaration(
                period_start_date_inclusive=start,
                period_end_date_exclusive=end,
                _revenue_ex_vat=None,  # will be computed
                _revenue_vat=None,
                _reverse_charge_outside_eu_ex_vat=None,
                _reverse_charge_outside_eu_vat=None,
                _reverse_charge_inside_eu_ex_vat=None,
                _reverse_charge_inside_eu_vat=None,
                _input_vat=None,
                _id=_id,
            )
        )
    return declarations


def _compute_vat_fields(data: Administracli) -> None:
    """Recalculate all computed VAT declaration fields from invoices."""
    from datetime import datetime

    def _to_date(d):
        if isinstance(d, datetime):
            return d.date()
        return d

    for decl in data.vat_declarations:
        start = _to_date(decl.period_start_date_inclusive)
        end = _to_date(decl.period_end_date_exclusive)

        revenue_ex_vat = Decimal(0)
        revenue_vat = Decimal(0)

        for inv in data.outgoing_invoices:
            inv_date = _to_date(inv.date)
            if inv_date is None or not (start <= inv_date < end):
                continue
            rate = inv.vat_rate
            ex_vat = inv.amount / (1 + rate) if rate else inv.amount
            revenue_ex_vat += ex_vat
            revenue_vat += inv.amount - ex_vat

        rc_outside_eu_ex = Decimal(0)
        rc_outside_eu_vat = Decimal(0)
        rc_inside_eu_ex = Decimal(0)
        rc_inside_eu_vat = Decimal(0)
        input_vat = Decimal(0)

        for inv in data.incoming_invoices:
            inv_date = _to_date(inv.date)
            if inv_date is None or not (start <= inv_date < end):
                continue

            if inv.vat_rate_abroad_from_outside_eu is not None:
                rate = inv.vat_rate_abroad_from_outside_eu
                # Reverse-charge: invoice amount IS the ex-VAT amount
                ex_vat = inv.amount
                vat = inv.amount * rate if rate else Decimal(0)
                rc_outside_eu_ex += ex_vat
                rc_outside_eu_vat += vat
                input_vat += vat  # reverse-charge VAT is deductible
            elif inv.vat_rate_abroad_from_inside_eu is not None:
                rate = inv.vat_rate_abroad_from_inside_eu
                # Reverse-charge: invoice amount IS the ex-VAT amount
                ex_vat = inv.amount
                vat = inv.amount * rate if rate else Decimal(0)
                rc_inside_eu_ex += ex_vat
                rc_inside_eu_vat += vat
                input_vat += vat  # reverse-charge VAT is deductible
            else:
                # Domestic purchase
                rate = inv.vat_rate
                if rate:
                    vat = inv.amount - inv.amount / (1 + rate)
                    input_vat += vat

        q = Decimal("1")  # Dutch VAT declaration rounds to whole euros
        decl._revenue_ex_vat = revenue_ex_vat.quantize(q)
        decl._revenue_vat = revenue_vat.quantize(q)
        decl._reverse_charge_outside_eu_ex_vat = rc_outside_eu_ex.quantize(q)
        decl._reverse_charge_outside_eu_vat = rc_outside_eu_vat.quantize(q)
        decl._reverse_charge_inside_eu_ex_vat = rc_inside_eu_ex.quantize(q)
        decl._reverse_charge_inside_eu_vat = rc_inside_eu_vat.quantize(q)
        decl._input_vat = input_vat.quantize(q)


def save_workbook(path: str, data: Administracli) -> None:
    """Write the Administracli data back to the Excel workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _write_transactions(wb, data.transactions)
    _write_incoming_invoices(wb, data.incoming_invoices)
    _write_outgoing_invoices(wb, data.outgoing_invoices)
    _write_vat_declarations(wb, data.vat_declarations)
    _write_settings(wb, data)

    wb.save(path)


def _write_transactions(wb: openpyxl.Workbook, transactions: list[Transaction]) -> None:
    ws = wb.create_sheet(title=TRANSACTIONS_SHEET)
    for col_idx, col_name in enumerate(TRANSACTION_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    field_map = {
        "Date": "date",
        "Amount": "amount",
        "Bank Account": "bank_account",
        "Description": "description",
        "_id": "_id",
        "_incoming_invoice_id": "_incoming_invoice_id",
        "_outgoing_invoice_id": "_outgoing_invoice_id",
        "_vat_declaration_id": "_vat_declaration_id",
        "_category": "_category",
    }

    for row_idx, txn in enumerate(transactions, start=2):
        for col_idx, col_name in enumerate(TRANSACTION_COLUMNS, start=1):
            attr = field_map[col_name]
            value = getattr(txn, attr)
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)


def _write_incoming_invoices(wb: openpyxl.Workbook, invoices: list[IncomingInvoice]) -> None:
    ws = wb.create_sheet(title=INCOMING_INVOICES_SHEET)
    columns = INCOMING_INVOICE_COLUMNS
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    field_map = {
        "Date": "date",
        "Amount": "amount",
        "Counterparty": "counterparty",
        "vat_rate": "vat_rate",
        "vat_rate_abroad_from_outside_eu": "vat_rate_abroad_from_outside_eu",
        "vat_rate_abroad_from_inside_eu": "vat_rate_abroad_from_inside_eu",
        "_id": "_id",
    }

    for row_idx, inv in enumerate(invoices, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = getattr(inv, field_map[col_name])
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)


def _write_outgoing_invoices(wb: openpyxl.Workbook, invoices: list[OutgoingInvoice]) -> None:
    ws = wb.create_sheet(title=OUTGOING_INVOICES_SHEET)
    columns = OUTGOING_INVOICE_COLUMNS
    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    field_map = {
        "Date": "date",
        "Amount": "amount",
        "Counterparty": "counterparty",
        "vat_rate": "vat_rate",
        "_id": "_id",
    }

    for row_idx, inv in enumerate(invoices, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            value = getattr(inv, field_map[col_name])
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)


def _write_vat_declarations(wb: openpyxl.Workbook, declarations: list[VATDeclaration]) -> None:
    ws = wb.create_sheet(title=VAT_DECLARATIONS_SHEET)
    for col_idx, col_name in enumerate(VAT_DECLARATION_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    field_map = {
        "period_start_date_inclusive": "period_start_date_inclusive",
        "period_end_date_exclusive": "period_end_date_exclusive",
        "_revenue_ex_vat": "_revenue_ex_vat",
        "_revenue_vat": "_revenue_vat",
        "_reverse_charge_outside_eu_ex_vat": "_reverse_charge_outside_eu_ex_vat",
        "_reverse_charge_outside_eu_vat": "_reverse_charge_outside_eu_vat",
        "_reverse_charge_inside_eu_ex_vat": "_reverse_charge_inside_eu_ex_vat",
        "_reverse_charge_inside_eu_vat": "_reverse_charge_inside_eu_vat",
        "_input_vat": "_input_vat",
        "_id": "_id",
    }

    for row_idx, decl in enumerate(declarations, start=2):
        for col_idx, col_name in enumerate(VAT_DECLARATION_COLUMNS, start=1):
            attr = field_map[col_name]
            value = getattr(decl, attr)
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)


def _write_settings(wb: openpyxl.Workbook, data: Administracli) -> None:
    ws = wb.create_sheet(title=SETTINGS_SHEET)
    ws.cell(row=1, column=1, value="Key")
    ws.cell(row=1, column=2, value="Value")
    ws.cell(row=2, column=1, value="cit_amount")
    ws.cell(row=2, column=2, value=float(data.cit_amount) if data.cit_amount is not None else None)

